import logging
import os
import time

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import WorkshopRegistration
from core.utils.sheets_utils import (
    append_workshop_to_excel_backup,
    append_workshop_to_google_sheet,
    build_workshop_data_row,
    build_workshop_header_row,
    get_existing_workshop_ids,
    get_sheets_service,
)


logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = "Backfill or rebuild Workshop registrations in Google Sheets and/or Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "--target",
            choices=["google", "excel", "both"],
            default="google",
            help="Where to write Workshop data.",
        )
        parser.add_argument(
            "--rebuild",
            action="store_true",
            help="Clear the target sheet and rebuild from DB.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="List records that would be processed without writing.",
        )
        parser.add_argument(
            "--sleep",
            type=float,
            default=1.0,
            help="Delay (seconds) between Google Sheets writes when backfilling.",
        )

    def handle(self, *args, **options):
        target = options.get("target", "google")
        rebuild = options.get("rebuild", False)
        dry_run = options.get("dry_run", False)
        sleep_seconds = options.get("sleep", 1.0)

        queryset = WorkshopRegistration.objects.order_by("id")
        total = queryset.count()

        self.stdout.write(f"Found {total} workshop registrations")

        if total == 0:
            return

        if rebuild:
            if dry_run:
                self.stdout.write("Dry run: rebuild skipped")
                return

            header_row = build_workshop_header_row()
            data_rows = [build_workshop_data_row(obj) for obj in queryset]

            if target in ("excel", "both"):
                self._rebuild_excel(header_row, data_rows)

            if target in ("google", "both"):
                self._rebuild_google(header_row, data_rows)

            return

        if target in ("google", "both"):
            self._backfill_google(queryset, total, dry_run, sleep_seconds)

        if target in ("excel", "both"):
            self._backfill_excel(queryset, total, dry_run)

    def _backfill_google(self, queryset, total, dry_run, sleep_seconds):
        existing_ids = set()
        try:
            service = get_sheets_service()
            existing_ids = get_existing_workshop_ids(service)
        except Exception as e:
            self.stderr.write(f"Failed to read existing Workshop IDs: {e}")

        added_count = 0
        updated_count = 0
        failed_count = 0

        for index, registration in enumerate(queryset, start=1):
            registration_id = str(registration.id)
            was_existing = registration_id in existing_ids

            if dry_run:
                action = "UPDATED" if was_existing else "ADDED"
                self.stdout.write(
                    f"[ {index} / {total} ] ID:{registration_id} {registration.email} -> {action} (dry-run)"
                )
                if was_existing:
                    updated_count += 1
                else:
                    added_count += 1
                continue

            try:
                result = append_workshop_to_google_sheet(registration)
                if result:
                    action = "UPDATED" if was_existing else "ADDED"
                    self.stdout.write(
                        f"[ {index} / {total} ] ID:{registration_id} {registration.email} -> {action}"
                    )
                    if was_existing:
                        updated_count += 1
                    else:
                        added_count += 1
                        existing_ids.add(registration_id)
                else:
                    failed_count += 1
                    self.stdout.write(
                        f"[ {index} / {total} ] ID:{registration_id} {registration.email} -> FAILED"
                    )
            except Exception as e:
                failed_count += 1
                self.stderr.write(
                    f"[ {index} / {total} ] ID:{registration_id} {registration.email} -> FAILED ({e})"
                )

            if sleep_seconds and sleep_seconds > 0:
                time.sleep(sleep_seconds)

        if dry_run:
            self.stdout.write(
                f"Complete: {added_count} added, {updated_count} updated, {failed_count} failed (dry-run)"
            )
        else:
            self.stdout.write(
                f"Complete: {added_count} added, {updated_count} updated, {failed_count} failed"
            )

        self.stdout.write("Check logs/sheets_errors.log for failures")

    def _backfill_excel(self, queryset, total, dry_run):
        if dry_run:
            self.stdout.write("Dry run: Excel backfill skipped")
            return

        processed = 0
        failed = 0

        for index, registration in enumerate(queryset, start=1):
            try:
                result = append_workshop_to_excel_backup(registration)
                if result:
                    processed += 1
                    self.stdout.write(
                        f"[ {index} / {total} ] ID:{registration.id} {registration.email} -> WRITTEN"
                    )
                else:
                    failed += 1
                    self.stdout.write(
                        f"[ {index} / {total} ] ID:{registration.id} {registration.email} -> FAILED"
                    )
            except Exception as e:
                failed += 1
                self.stderr.write(
                    f"[ {index} / {total} ] ID:{registration.id} {registration.email} -> FAILED ({e})"
                )

        self.stdout.write(f"Complete: {processed} written, {failed} failed")

    def _rebuild_excel(self, header_row, data_rows):
        exports_dir = settings.BASE_DIR / "exports"
        os.makedirs(exports_dir, exist_ok=True)
        excel_path = exports_dir / "approved_registrations.xlsx"

        if excel_path.exists():
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()

        if "Workshop" in wb.sheetnames:
            wb.remove(wb["Workshop"])
        ws = wb.create_sheet(title="Workshop")

        ws.append(header_row)

        payment_columns = []
        for idx, header in enumerate(header_row, start=1):
            lowered = str(header).lower()
            if "transaction" in lowered or "payment" in lowered or "fee" in lowered:
                payment_columns.append(idx)

        yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
        bold = Font(bold=True)

        for row in data_rows:
            ws.append(row)

        if payment_columns:
            for col_idx in payment_columns:
                header_cell = ws.cell(row=1, column=col_idx)
                header_cell.fill = yellow
                header_cell.font = bold

            for row_idx in range(2, ws.max_row + 1):
                for col_idx in payment_columns:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = yellow
                    cell.font = bold

        wb.save(excel_path)
        self.stdout.write(self.style.SUCCESS(f"Excel Workshop sheet rebuilt at {excel_path}"))

    def _rebuild_google(self, header_row, data_rows):
        if not settings.GOOGLE_SHEETS_ID:
            self.stdout.write(
                self.style.WARNING("GOOGLE_SHEETS_ID not configured; skipping Google Sheets")
            )
            return

        service = get_sheets_service()
        sheet_id = settings.GOOGLE_SHEETS_ID
        end_col = get_column_letter(len(header_row))
        clear_range = "Workshop!A1:ZZ"

        service.spreadsheets().values().clear(
            spreadsheetId=sheet_id,
            range=clear_range,
            body={},
        ).execute()

        rows = [header_row] + data_rows
        batch_size = 500
        start_row = 1

        for idx in range(0, len(rows), batch_size):
            chunk = rows[idx:idx + batch_size]
            end_row = start_row + len(chunk) - 1
            range_name = f"Workshop!A{start_row}:{end_col}{end_row}"

            service.spreadsheets().values().update(
                spreadsheetId=sheet_id,
                range=range_name,
                valueInputOption="USER_ENTERED",
                body={"values": chunk},
            ).execute()

            start_row = end_row + 1

        self.stdout.write(self.style.SUCCESS("Google Sheets Workshop tab rebuilt"))
