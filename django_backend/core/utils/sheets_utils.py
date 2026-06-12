# core/utils/sheets_utils.py
"""
Google Sheets integration utility for exporting approved registrations.
Implements retry/backoff, serialized sync (via signals.py queue workers),
idempotency safeguards, and failure-visible fallback logging.
"""
import logging
import os
import time
from datetime import datetime
from pathlib import Path

from django.conf import settings
from google.auth.transport.requests import Request
from google.oauth2.service_account import Credentials
from googleapiclient import discovery
from googleapiclient.errors import HttpError
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Retry helper
# ---------------------------------------------------------------------------

# Transient HTTP status codes worth retrying (rate-limit + server errors).
# 4xx auth/permission errors (401, 403, 404) are NOT retried — fail fast.
_RETRYABLE_STATUS = {429, 500, 502, 503}
_RETRY_DELAYS = (1, 3, 8)  # seconds between attempts 1→2, 2→3, 3→fail


def _call_with_retry(api_request):
    """
    Execute a Google API request object with exponential backoff.

    Args:
        api_request: A googleapiclient request object (not yet .execute()'d)

    Returns:
        The API response dict.

    Raises:
        HttpError: On non-retryable errors or after all retries exhausted.
        Exception:  On non-HTTP errors.
    """
    last_exc = None
    for attempt, delay in enumerate((_RETRY_DELAYS + (None,)), start=1):
        try:
            return api_request.execute()
        except HttpError as exc:
            status = exc.resp.status if exc.resp else None
            if status not in _RETRYABLE_STATUS:
                # Auth / permission / not-found errors: no point retrying.
                logger.error(
                    "Google Sheets API non-retryable error (HTTP %s): %s",
                    status, exc,
                )
                raise
            last_exc = exc
            if delay is None:
                break  # retries exhausted
            logger.warning(
                "Google Sheets API transient error (HTTP %s), attempt %d — retrying in %ds",
                status, attempt, delay,
            )
            time.sleep(delay)

    raise last_exc


# ---------------------------------------------------------------------------
# Authentication
# ---------------------------------------------------------------------------

def get_sheets_service():
    """
    Authenticates using credentials.json and returns Google Sheets API service.

    Returns:
        googleapiclient.discovery.Resource: Authenticated Sheets API service

    Raises:
        FileNotFoundError: If credentials.json is not found
        Exception: If authentication fails
    """
    credentials_path = settings.GOOGLE_SHEETS_CREDENTIALS_JSON

    if not os.path.exists(credentials_path):
        # Do NOT log the full path — it may reveal deployment layout.
        logger.error("Google Sheets credentials file not found")
        raise FileNotFoundError("Google Sheets credentials not found")

    try:
        credentials = Credentials.from_service_account_file(
            credentials_path,
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        service = discovery.build("sheets", "v4", credentials=credentials)
        logger.debug("Google Sheets service authenticated successfully")
        return service
    except Exception as exc:
        logger.error("Failed to authenticate Google Sheets: %s", type(exc).__name__)
        raise


# ---------------------------------------------------------------------------
# Registration (Approvals tab) helpers
# ---------------------------------------------------------------------------

def get_existing_registration_ids(service):
    """
    Reads column A (registration IDs) from the Approvals sheet tab.

    Returns:
        set[str]: IDs already in the sheet (empty set on read failure).
    """
    try:
        sheet_id = settings.GOOGLE_SHEETS_ID
        range_name = settings.GOOGLE_SHEETS_RANGE
        result = _call_with_retry(
            service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range=f"{range_name.split('!')[0]}!A:A",
            )
        )
        values = result.get("values", [])
        existing_ids = {str(row[0]) for row in values[1:] if row}
        logger.debug("Found %d existing registration IDs in sheet", len(existing_ids))
        return existing_ids
    except Exception as exc:
        logger.error("Failed to read existing registration IDs: %s", type(exc).__name__)
        return set()


def build_header_row():
    """Dynamically builds header row from Registration model fields."""
    from core.models import Registration
    field_names = [field.name for field in Registration._meta.fields]
    headers = field_names + ["Registration Status"]
    logger.debug("Built header row with %d columns", len(headers))
    return headers


def build_data_row(registration_instance):
    """
    Dynamically builds data row from a Registration instance.
    NOTE: intentionally does NOT log field values to avoid PII leakage.
    """
    from core.models import Registration
    field_names = [field.name for field in Registration._meta.fields]
    data_row = []
    for field_name in field_names:
        value = getattr(registration_instance, field_name, None)
        if value is None:
            data_row.append("")
        elif isinstance(value, datetime):
            data_row.append(value.isoformat())
        elif hasattr(value, "id"):  # ForeignKey
            data_row.append(str(value.id))
        else:
            data_row.append(str(value))
    data_row.append(registration_instance.get_status_display())
    logger.debug("Built data row for registration ID %s", registration_instance.id)
    return data_row


def ensure_header_exists(service, sheet_id, range_name):
    """
    Writes header row to row 1 if the sheet is empty.

    Returns:
        bool: True if header exists or was written successfully.
    """
    try:
        result = _call_with_retry(
            service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range=f"{range_name.split('!')[0]}!A1:ZZ1",
            )
        )
        values = result.get("values", [])
        if not values or len(values[0]) == 0:
            header_row = build_header_row()
            _call_with_retry(
                service.spreadsheets().values().update(
                    spreadsheetId=sheet_id,
                    range=f"{range_name.split('!')[0]}!A1:{get_column_letter(len(header_row))}1",
                    valueInputOption="USER_ENTERED",
                    body={"values": [header_row]},
                )
            )
            logger.info("Header row written to Google Sheet")
        else:
            logger.debug("Header row already exists in Google Sheet")
        return True
    except Exception as exc:
        logger.error("Failed to ensure header exists: %s", type(exc).__name__)
        return False


def update_registration_in_sheet(registration_instance, service):
    """
    Finds the row for registration_instance.id in the Approvals sheet and
    updates it in-place.

    Returns:
        bool: True if the row was found and updated; False if not found.
    """
    try:
        sheet_id = settings.GOOGLE_SHEETS_ID
        range_name = settings.GOOGLE_SHEETS_RANGE
        sheet_name = range_name.split("!")[0]
        registration_id_str = str(registration_instance.id)

        result = _call_with_retry(
            service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range=f"{sheet_name}!A:A",
            )
        )
        values = result.get("values", [])

        row_index = None
        for idx, row in enumerate(values[1:], start=2):
            if row and row[0] and str(row[0]) == registration_id_str:
                row_index = idx
                break

        if row_index is None:
            logger.warning(
                "Registration %s not found in sheet for update", registration_id_str
            )
            return False

        data_row = build_data_row(registration_instance)
        header_row = build_header_row()
        end_col = get_column_letter(len(header_row))
        _call_with_retry(
            service.spreadsheets().values().update(
                spreadsheetId=sheet_id,
                range=f"{sheet_name}!A{row_index}:{end_col}{row_index}",
                valueInputOption="USER_ENTERED",
                body={"values": [data_row]},
            )
        )
        logger.info(
            "Registration %s updated in Google Sheet at row %d (status=%s)",
            registration_id_str, row_index, registration_instance.status,
        )
        return True
    except Exception as exc:
        logger.error(
            "Failed to update registration %s in Google Sheet: %s",
            registration_instance.id, type(exc).__name__,
        )
        return False


def append_approved_user_to_sheet(registration_instance):
    """
    Main export function: syncs a Registration to the Approvals sheet tab.

    - If the ID already exists in the sheet: updates the row in-place.
    - Otherwise: appends a new row (after verifying the ID is still absent,
      to guard against a TOCTOU duplicate inside the same serialized job).
    - On any Google Sheets failure: falls back to the local Excel backup and
      logs at ERROR level with a SHEET_SYNC_FALLBACK marker so ops can alert.

    Returns:
        bool: True if data landed somewhere (sheet OR Excel); False only if
              both destinations failed.
    """
    reg_id = registration_instance.id
    logger.info(
        "Starting export for registration ID %s (status=%s)",
        reg_id, registration_instance.status,
    )

    # Accept every status that Registration.STATUS_CHOICES defines — derived
    # at call time so adding a new status in models.py is automatically covered.
    from core.models import Registration as _Reg
    valid_statuses = {choice[0] for choice in _Reg.STATUS_CHOICES}
    if registration_instance.status not in valid_statuses:
        logger.warning(
            "Registration %s has unknown status '%s'; skipping export",
            reg_id, registration_instance.status,
        )
        return False

    try:
        service = get_sheets_service()
        sheet_id = settings.GOOGLE_SHEETS_ID
        range_name = settings.GOOGLE_SHEETS_RANGE

        existing_ids = get_existing_registration_ids(service)
        registration_id_str = str(reg_id)

        if registration_id_str in existing_ids:
            logger.info(
                "Registration %s already in sheet — updating (status=%s)",
                registration_id_str, registration_instance.status,
            )
            if update_registration_in_sheet(registration_instance, service):
                return True
            # Update returned False (row disappeared between reads).
            # Fall through to the append path with a re-check below.
            logger.warning(
                "Update returned False for registration %s; will re-check and append",
                registration_id_str,
            )

        # Ensure header, then re-verify ID is absent before appending to
        # eliminate any residual TOCTOU window within this serialized job.
        if not ensure_header_exists(service, sheet_id, range_name):
            raise Exception("Failed to ensure header row exists")

        # Re-check: another job (e.g. backfill) might have appended this ID
        # between our first get_existing_registration_ids() call and now.
        re_check_ids = get_existing_registration_ids(service)
        if registration_id_str in re_check_ids:
            logger.info(
                "Registration %s appeared in sheet during re-check — updating instead of appending",
                registration_id_str,
            )
            update_registration_in_sheet(registration_instance, service)
            return True

        data_row = build_data_row(registration_instance)
        _call_with_retry(
            service.spreadsheets().values().append(
                spreadsheetId=sheet_id,
                range=range_name,
                valueInputOption="USER_ENTERED",
                body={"values": [data_row]},
            )
        )
        logger.info(
            "Registration %s appended to Google Sheet (status=%s)",
            registration_id_str, registration_instance.status,
        )
        return True

    except FileNotFoundError:
        logger.error(
            "SHEET_SYNC_FALLBACK registration_id=%s model=Registration reason=credentials_not_found",
            reg_id,
        )
        return append_to_excel_backup(registration_instance)

    except Exception as exc:
        logger.error(
            "SHEET_SYNC_FALLBACK registration_id=%s model=Registration reason=%s",
            reg_id, type(exc).__name__,
        )
        if append_to_excel_backup(registration_instance):
            logger.info(
                "Registration %s written to Excel backup after Google Sheet failure", reg_id
            )
            return True
        logger.error(
            "Both Google Sheet AND Excel backup failed for registration %s", reg_id
        )
        return False


# ---------------------------------------------------------------------------
# Registration Excel backup
# ---------------------------------------------------------------------------

def append_to_excel_backup(registration_instance):
    """
    Appends or updates the registration in the local Excel backup file.

    Returns:
        bool: True if successful.
    """
    try:
        exports_dir = settings.BASE_DIR / "exports"
        os.makedirs(exports_dir, exist_ok=True)
        excel_path = exports_dir / "approved_registrations.xlsx"

        data_row = build_data_row(registration_instance)
        registration_id = str(registration_instance.id)

        if excel_path.exists():
            wb = load_workbook(excel_path)
            ws = wb.active
            existing_row_num = None
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1), start=2
            ):
                if row[0].value and str(row[0].value) == registration_id:
                    existing_row_num = row_idx
                    break

            if existing_row_num:
                for col_idx, value in enumerate(data_row, start=1):
                    ws.cell(row=existing_row_num, column=col_idx, value=value)
                wb.save(excel_path)
                logger.info(
                    "Registration %s updated in Excel backup (status=%s)",
                    registration_id, registration_instance.status,
                )
                return True
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(build_header_row())

        ws = wb.active
        ws.append(data_row)
        wb.save(excel_path)
        logger.info(
            "Registration %s appended to Excel backup (status=%s)",
            registration_id, registration_instance.status,
        )
        return True

    except Exception as exc:
        logger.error(
            "Failed to write Excel backup for registration %s: %s",
            registration_instance.id, type(exc).__name__,
        )
        return False


# ---------------------------------------------------------------------------
# Workshop tab helpers
# ---------------------------------------------------------------------------

def get_workshop_sheet_range():
    """Returns the range string for the Workshop tab."""
    return getattr(settings, "GOOGLE_SHEETS_WORKSHOP_RANGE", "Workshop!A1")


def get_existing_workshop_ids(service):
    """
    Reads column A (workshop registration IDs) from the Workshop tab.

    Returns:
        set[str]: IDs already in the sheet (empty set on read failure).
    """
    try:
        sheet_id = settings.GOOGLE_SHEETS_ID
        result = _call_with_retry(
            service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range="Workshop!A:A",
            )
        )
        values = result.get("values", [])
        existing_ids = {str(row[0]) for row in values[1:] if row}
        logger.debug("Found %d existing workshop IDs in sheet", len(existing_ids))
        return existing_ids
    except Exception as exc:
        logger.error("Failed to read existing workshop IDs: %s", type(exc).__name__)
        return set()


def build_workshop_header_row():
    """Dynamically builds header row from WorkshopRegistration model fields."""
    from core.models import WorkshopRegistration

    fields = list(WorkshopRegistration._meta.fields)
    field_names = [field.name for field in fields]

    def is_payment_field(name):
        lowered = name.lower()
        return "transaction" in lowered or "payment" in lowered or "fee" in lowered

    payment_field_names = [name for name in field_names if is_payment_field(name)]
    non_payment_field_names = [name for name in field_names if name not in payment_field_names]

    ordered_names = []
    for preferred in ["id", "full_name", "email", "workshop_id", "workshop_title", "status", "created_at"]:
        if preferred in non_payment_field_names:
            ordered_names.append(preferred)
    for name in non_payment_field_names:
        if name not in ordered_names:
            ordered_names.append(name)
    ordered_names.extend(payment_field_names)

    header_row = []
    for name in ordered_names:
        field = next((f for f in fields if f.name == name), None)
        header_row.append(
            field.verbose_name.replace("_", " ").title() if field else name
        )
    header_row.append("Last Updated")

    logger.debug("Built workshop header row with %d columns", len(header_row))
    return header_row


def build_workshop_data_row(workshop_instance):
    """Builds data row for workshop registration matching header order."""
    from core.models import WorkshopRegistration

    fields = list(WorkshopRegistration._meta.fields)
    field_names = [field.name for field in fields]

    def is_payment_field(name):
        lowered = name.lower()
        return "transaction" in lowered or "payment" in lowered or "fee" in lowered

    payment_field_names = [name for name in field_names if is_payment_field(name)]
    non_payment_field_names = [name for name in field_names if name not in payment_field_names]

    ordered_names = []
    for preferred in ["id", "full_name", "email", "workshop_id", "workshop_title", "status", "created_at"]:
        if preferred in non_payment_field_names:
            ordered_names.append(preferred)
    for name in non_payment_field_names:
        if name not in ordered_names:
            ordered_names.append(name)
    ordered_names.extend(payment_field_names)

    data_row = []
    for name in ordered_names:
        value = getattr(workshop_instance, name, None)
        lowered = name.lower()

        if "transaction_id" in lowered:
            data_row.append(str(value) if str(value).strip() else "PENDING")
            continue
        if "screenshot" in lowered:
            data_row.append("Uploaded ✓" if value else "Not Uploaded ✗")
            continue
        if lowered == "status" or "payment_status" in lowered:
            status_value = str(value).strip()
            data_row.append(f"★ {status_value}" if status_value else "★ PENDING")
            continue
        if value is None:
            data_row.append("")
            continue
        if isinstance(value, datetime):
            data_row.append(value.isoformat())
        elif hasattr(value, "id"):
            data_row.append(str(value.id))
        else:
            data_row.append(str(value))

    data_row.append(datetime.now().strftime("%d-%m-%Y %H:%M"))
    logger.debug("Built workshop data row for ID %s", workshop_instance.id)
    return data_row


def ensure_workshop_header_exists(service):
    """
    Ensures the Workshop tab has a header row.

    Returns:
        tuple[bool, bool]: (success, header_was_created)
    """
    try:
        result = _call_with_retry(
            service.spreadsheets().values().get(
                spreadsheetId=settings.GOOGLE_SHEETS_ID,
                range="Workshop!A1:ZZ1",
            )
        )
        values = result.get("values", [])
        if not values or len(values[0]) == 0:
            header_row = build_workshop_header_row()
            _call_with_retry(
                service.spreadsheets().values().update(
                    spreadsheetId=settings.GOOGLE_SHEETS_ID,
                    range="Workshop!A1",
                    valueInputOption="USER_ENTERED",
                    body={"values": [header_row]},
                )
            )
            logger.info("Workshop header row written to Google Sheet")
            return True, True
        logger.debug("Workshop header row already exists")
        return True, False
    except Exception as exc:
        logger.error("Failed to ensure workshop header exists: %s", type(exc).__name__)
        return False, False


def append_workshop_to_google_sheet(workshop_instance):
    """
    Main export function: syncs a WorkshopRegistration to the Workshop sheet tab.

    - If the ID already exists: updates the row in-place.
    - Otherwise: re-verifies absence, then appends (idempotency safeguard).
    - On any Google Sheets failure: falls back to local Excel backup and logs
      at ERROR level with a SHEET_SYNC_FALLBACK marker.

    Returns:
        bool: True if data landed somewhere (sheet OR Excel).
    """
    ws_id = workshop_instance.id
    logger.info(
        "Starting Workshop export for ID %s (status=%s)",
        ws_id, workshop_instance.status,
    )

    try:
        service = get_sheets_service()
        sheet_id = settings.GOOGLE_SHEETS_ID

        header_ok, header_created = ensure_workshop_header_exists(service)
        if not header_ok:
            raise Exception("Failed to ensure Workshop header row exists")

        existing_ids = get_existing_workshop_ids(service)
        registration_id_str = str(ws_id)
        data_row = build_workshop_data_row(workshop_instance)

        if registration_id_str in existing_ids:
            # Re-fetch column A to find exact row index.
            result = _call_with_retry(
                service.spreadsheets().values().get(
                    spreadsheetId=sheet_id,
                    range="Workshop!A:A",
                )
            )
            values = result.get("values", [])
            row_index = None
            for idx, row in enumerate(values[1:], start=2):
                if row and row[0] and str(row[0]) == registration_id_str:
                    row_index = idx
                    break

            if row_index:
                _call_with_retry(
                    service.spreadsheets().values().update(
                        spreadsheetId=sheet_id,
                        range=f"Workshop!A{row_index}",
                        valueInputOption="USER_ENTERED",
                        body={"values": [data_row]},
                    )
                )
                logger.info(
                    "Workshop registration %s updated at row %d", registration_id_str, row_index
                )
            else:
                # ID was in existing_ids but row not found (header shift / deletion).
                # Guard: re-check before appending.
                re_check = get_existing_workshop_ids(service)
                if registration_id_str not in re_check:
                    _call_with_retry(
                        service.spreadsheets().values().append(
                            spreadsheetId=sheet_id,
                            range=get_workshop_sheet_range(),
                            valueInputOption="USER_ENTERED",
                            insertDataOption="INSERT_ROWS",
                            body={"values": [data_row]},
                        )
                    )
                    logger.info(
                        "Workshop registration %s appended after row-not-found re-check",
                        registration_id_str,
                    )
                else:
                    logger.warning(
                        "Workshop registration %s: row index unclear after re-check; skipping append to avoid duplicate",
                        registration_id_str,
                    )
        else:
            # Idempotency safeguard: re-verify absence before appending.
            re_check = get_existing_workshop_ids(service)
            if registration_id_str in re_check:
                logger.info(
                    "Workshop registration %s appeared during re-check — updating instead",
                    registration_id_str,
                )
                return append_workshop_to_google_sheet(workshop_instance)

            _call_with_retry(
                service.spreadsheets().values().append(
                    spreadsheetId=sheet_id,
                    range=get_workshop_sheet_range(),
                    valueInputOption="USER_ENTERED",
                    insertDataOption="INSERT_ROWS",
                    body={"values": [data_row]},
                )
            )
            logger.info(
                "Workshop registration %s appended to Google Sheet", registration_id_str
            )

        # Apply conditional formatting when header was just created.
        if header_created:
            _apply_workshop_payment_formatting(service, sheet_id, workshop_instance)

        logger.info("Workshop registration %s export completed", ws_id)
        return True

    except Exception as exc:
        logger.error(
            "SHEET_SYNC_FALLBACK registration_id=%s model=WorkshopRegistration reason=%s",
            ws_id, type(exc).__name__,
        )
        if append_workshop_to_excel_backup(workshop_instance):
            logger.info(
                "Workshop registration %s written to Excel backup after Google Sheet failure", ws_id
            )
            return True
        logger.error(
            "Both Google Sheet AND Excel backup failed for workshop registration %s", ws_id
        )
        return False


def _apply_workshop_payment_formatting(service, sheet_id, workshop_instance):
    """Applies yellow header formatting to payment columns on first header creation."""
    try:
        metadata = _call_with_retry(
            service.spreadsheets().get(
                spreadsheetId=sheet_id,
                fields="sheets(properties(sheetId,title),conditionalFormats)",
            )
        )
        workshop_sheet = next(
            (s for s in metadata.get("sheets", [])
             if s.get("properties", {}).get("title") == "Workshop"),
            None,
        )
        if not workshop_sheet or workshop_sheet.get("conditionalFormats"):
            return

        fields = [f.name for f in workshop_instance._meta.fields]
        payment_names = [
            name for name in fields
            if "transaction" in name.lower() or "payment" in name.lower() or "fee" in name.lower()
        ]
        ordered_fields = []
        for preferred in ["id", "full_name", "email", "workshop_id", "workshop_title", "status", "created_at"]:
            if preferred in fields:
                ordered_fields.append(preferred)
        for name in fields:
            if name not in ordered_fields and name not in payment_names:
                ordered_fields.append(name)
        ordered_fields.extend(payment_names)

        payment_indices = [idx for idx, name in enumerate(ordered_fields) if name in payment_names]
        if not payment_indices:
            return

        payment_start_index = min(payment_indices)
        payment_end_index = max(payment_indices) + 1
        ws_sheet_id = workshop_sheet.get("properties", {}).get("sheetId")

        request = {
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [{
                        "sheetId": ws_sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": 1,
                        "startColumnIndex": payment_start_index,
                        "endColumnIndex": payment_end_index,
                    }],
                    "booleanRule": {
                        "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=TRUE"}],
                        },
                        "format": {"backgroundColor": {"red": 1.0, "green": 1.0, "blue": 0.0}},
                    },
                },
                "index": 0,
            }
        }
        _call_with_retry(
            service.spreadsheets().batchUpdate(
                spreadsheetId=sheet_id,
                body={"requests": [request]},
            )
        )
        logger.info("Applied Workshop header payment column formatting")
    except Exception as exc:
        logger.warning("Failed to apply Workshop header formatting: %s", type(exc).__name__)


# ---------------------------------------------------------------------------
# Workshop Excel backup
# ---------------------------------------------------------------------------

def append_workshop_to_excel_backup(workshop_instance):
    """
    Fallback export to Excel when Google Sheets is unavailable.

    Returns:
        bool: True if successful.
    """
    try:
        from openpyxl.styles import Font, PatternFill

        exports_dir = settings.BASE_DIR / "exports"
        os.makedirs(exports_dir, exist_ok=True)
        excel_path = exports_dir / "approved_registrations.xlsx"

        header_row = build_workshop_header_row()
        data_row = build_workshop_data_row(workshop_instance)
        registration_id = str(workshop_instance.id)

        if excel_path.exists():
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()

        if "Workshop" in wb.sheetnames:
            ws = wb["Workshop"]
        else:
            ws = wb.create_sheet(title="Workshop")

        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
            ws.append(header_row)

        payment_columns = [
            idx for idx, header in enumerate(header_row, start=1)
            if any(kw in str(header).lower() for kw in ("transaction", "payment", "fee"))
        ]
        yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
        bold = Font(bold=True)

        existing_row_num = None
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and str(cell_value) == registration_id:
                existing_row_num = row_idx
                break

        target_row = existing_row_num or (ws.max_row + 1)
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            if col_idx in payment_columns:
                cell.fill = yellow
                cell.font = bold

        for col_idx in payment_columns:
            hc = ws.cell(row=1, column=col_idx)
            hc.fill = yellow
            hc.font = bold

        wb.save(excel_path)
        action = "updated" if existing_row_num else "appended"
        logger.info("Workshop registration %s %s in Excel backup", registration_id, action)
        return True

    except Exception as exc:
        logger.error(
            "Failed to write Workshop Excel backup for %s: %s",
            workshop_instance.id, type(exc).__name__,
        )
        return False
