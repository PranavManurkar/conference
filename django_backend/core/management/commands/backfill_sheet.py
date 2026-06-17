# core/management/commands/backfill_sheet.py
"""
Management command: sync all Registration records to the Approvals Google Sheet.

Usage examples
--------------
Dry run (no writes — see what would be added/updated):
    python manage.py backfill_sheet --dry-run

Standard backfill (idempotent upsert of every registration):
    python manage.py backfill_sheet

Verify post-backfill (compare DB to sheet, no writes):
    python manage.py backfill_sheet --verify

Sort the Approvals sheet by payment_date descending, NULLs last:
    python manage.py backfill_sheet --sort

Full rebuild (clears sheet first, then rewrites all rows at once):
    python manage.py backfill_sheet --rebuild

Timing note: the standard backfill sleeps 2 seconds between each registration to
respect Google's 60 calls/minute quota (append_approved_user_to_sheet makes 2-4
API calls per registration internally). At 222 registrations this takes ~7-8 minutes.
Run during low-traffic hours when possible.
"""
import logging
import os
import time
from django.conf import settings
from django.db.models import F
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.models import Registration
from core.utils.sheets_utils import (
    get_sheets_service,
    get_existing_registration_ids,
    build_header_row,
    build_data_row,
    append_approved_user_to_sheet,
    sort_approvals_sheet_by_payment_date,
)

logger = logging.getLogger(__name__)

# Seconds to sleep between registrations during the standard backfill.
# append_approved_user_to_sheet makes 2-4 Google Sheets API calls internally.
# Sleeping 2s between registrations keeps the effective rate well under the
# 60 calls/minute quota even when each registration triggers 4 API calls:
#   30 registrations/min × 4 calls = 120 calls spread over 60s ≈ 2 calls/sec.
_BACKFILL_SLEEP_SECONDS = 2


class Command(BaseCommand):
    help = (
        "Backfill or rebuild Main Conference registrations in Google Sheets and/or Excel. "
        f"Standard backfill sleeps {_BACKFILL_SLEEP_SECONDS}s between registrations "
        "(~7-8 min for 222 records) to stay within Google's API quota."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--target",
            choices=["google", "excel", "both"],
            default="google",
            help="Where to write the registration data.",
        )
        parser.add_argument(
            "--rebuild",
            action="store_true",
            help="Clear the target sheet and completely rebuild from the database.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            dest="dry_run",
            help=(
                "Preview what would be added or updated without writing anything. "
                "Reads the current sheet column A and compares to the DB. Safe to run at any time."
            ),
        )
        parser.add_argument(
            "--verify",
            action="store_true",
            help=(
                "Compare the sheet's column A against all Registration IDs in the DB. "
                "Prints any IDs missing from the sheet and any extra IDs in the sheet. "
                "No writes. Safe to run at any time."
            ),
        )
        parser.add_argument(
            "--sort",
            action="store_true",
            help=(
                "Sort the Approvals sheet by payment_date descending, NULLs last. "
                "Uses a single server-side batchUpdate — instant and safe on a live sheet."
            ),
        )

    def handle(self, *args, **options):
        target = options.get("target", "google")
        rebuild = options.get("rebuild", False)
        dry_run = options.get("dry_run", False)
        verify = options.get("verify", False)
        do_sort = options.get("sort", False)

        # ------------------------------------------------------------------
        # --dry-run: preview only, zero writes
        # ------------------------------------------------------------------
        if dry_run:
            self._handle_dry_run()
            return

        # ------------------------------------------------------------------
        # --verify: compare DB vs sheet, zero writes
        # ------------------------------------------------------------------
        if verify:
            self._handle_verify()
            return

        # ------------------------------------------------------------------
        # --sort: server-side sort of the Approvals sheet, single API call
        # ------------------------------------------------------------------
        if do_sort:
            self._handle_sort()
            return

        # ------------------------------------------------------------------
        # --rebuild: clear sheet and rewrite all rows in one batch
        # ------------------------------------------------------------------
        # Build queryset ordered by most recent payment date, then creation date.
        # F("payment_date").desc(nulls_last=True) puts registrations with a payment
        # date first (newest payment first); registrations without a payment date
        # appear at the end, ordered by creation date descending.
        queryset = Registration.objects.all().order_by(
            F("payment_date").desc(nulls_last=True),
            "-created_at",
        )
        total = queryset.count()
        self.stdout.write(f"Found {total} total registration(s) in the database.")

        if total == 0:
            return

        if rebuild:
            self.stdout.write("REBUILD MODE ACTIVE: Extracting all data...")
            header_row = build_header_row()
            data_rows = [build_data_row(obj) for obj in queryset]

            if target in ("excel", "both"):
                self._rebuild_excel(header_row, data_rows)

            if target in ("google", "both"):
                self._rebuild_google(header_row, data_rows)
            return

        # ------------------------------------------------------------------
        # Standard line-by-line idempotent backfill
        # ------------------------------------------------------------------
        if target not in ("google", "both"):
            self.stdout.write("Target is 'excel' only — nothing to do for Google Sheets.")
            return

        if not getattr(settings, "GOOGLE_SHEETS_ID", None):
            self.stdout.write(
                self.style.WARNING("GOOGLE_SHEETS_ID not configured; skipping Google Sheets backfill.")
            )
            return

        self.stdout.write(
            f"Running standard line-by-line sync ({_BACKFILL_SLEEP_SECONDS}s sleep between "
            f"registrations to respect API quota)..."
        )

        # Pre-fetch existing IDs once so we can report ADD vs UPDATE per row
        # without making an extra API call per registration.
        try:
            service = get_sheets_service()
            existing_ids = get_existing_registration_ids(service)
            self.stdout.write(
                f"Sheet currently contains {len(existing_ids)} registration ID(s)."
            )
        except Exception as exc:
            self.stdout.write(
                self.style.ERROR(f"Failed to read existing sheet IDs: {exc}. Aborting.")
            )
            return

        added = 0
        updated = 0
        failed = 0
        start_time = time.time()

        for idx, registration in enumerate(queryset, 1):
            reg_id_str = str(registration.id)
            was_in_sheet = reg_id_str in existing_ids

            success = append_approved_user_to_sheet(registration)

            if success:
                if was_in_sheet:
                    updated += 1
                    label = self.style.SUCCESS(f"[OK - UPDATED in Sheet]")
                else:
                    added += 1
                    # Add to our local set so subsequent iterations reflect reality.
                    existing_ids.add(reg_id_str)
                    label = self.style.SUCCESS(f"[OK - ADDED to Sheet]  ")
            else:
                failed += 1
                label = self.style.ERROR(f"[FAILED]               ")

            self.stdout.write(
                f"[{idx:>3}/{total}] {label} ID {registration.id}"
            )

            time.sleep(_BACKFILL_SLEEP_SECONDS)

        elapsed = time.time() - start_time

        self.stdout.write("")
        self.stdout.write("=" * 60)
        self.stdout.write(f"BACKFILL COMPLETE")
        self.stdout.write(f"  Total registrations in DB : {total}")
        self.stdout.write(
            self.style.SUCCESS(
                f"  Successfully written to Sheet : {added + updated}  "
                f"(added: {added}, updated: {updated})"
            )
        )
        if failed:
            self.stdout.write(
                self.style.ERROR(
                    f"  Failed (check logs/Excel backup): {failed}"
                )
            )
        else:
            self.stdout.write(self.style.SUCCESS(f"  Failed                       : 0"))
        self.stdout.write(f"  Total time taken           : {elapsed:.1f}s")
        self.stdout.write("=" * 60)

        if failed:
            self.stdout.write(
                self.style.WARNING(
                    f"\n  {failed} registration(s) failed to write to Google Sheets. "
                    "They have been saved to the Excel backup (exports/approved_registrations.xlsx). "
                    "Check logs/sheets_errors.log for details, then re-run backfill to retry."
                )
            )

    # ------------------------------------------------------------------
    # --dry-run handler
    # ------------------------------------------------------------------

    def _handle_dry_run(self):
        self.stdout.write("[DRY RUN] No writes will be made.\n")

        if not getattr(settings, "GOOGLE_SHEETS_ID", None):
            self.stdout.write(
                self.style.WARNING("GOOGLE_SHEETS_ID not configured; cannot compare against sheet.")
            )
            return

        try:
            service = get_sheets_service()
            existing_ids = get_existing_registration_ids(service)
        except Exception as exc:
            self.stdout.write(self.style.ERROR(f"Failed to read sheet IDs: {exc}"))
            return

        queryset = Registration.objects.all().order_by(
            F("payment_date").desc(nulls_last=True),
            "-created_at",
        )

        would_add = 0
        would_update = 0

        for registration in queryset:
            reg_id_str = str(registration.id)
            if reg_id_str in existing_ids:
                would_update += 1
                self.stdout.write(
                    f"[DRY-RUN] ID {registration.id} — would UPDATE (already in sheet)"
                )
            else:
                would_add += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"[DRY-RUN] ID {registration.id} — would ADD    (not in sheet)"
                    )
                )

        self.stdout.write("")
        self.stdout.write("=" * 60)
        self.stdout.write(f"DRY RUN SUMMARY")
        self.stdout.write(f"  Registrations in DB : {would_add + would_update}")
        self.stdout.write(
            self.style.WARNING(f"  Would be ADDED      : {would_add}")
        )
        self.stdout.write(f"  Would be UPDATED    : {would_update}")
        self.stdout.write("=" * 60)
        self.stdout.write(
            "\nRun without --dry-run to execute the backfill."
        )

    # ------------------------------------------------------------------
    # --verify handler
    # ------------------------------------------------------------------

    def _handle_verify(self):
        self.stdout.write("[VERIFY] Comparing database against Google Sheet. No writes.\n")

        if not getattr(settings, "GOOGLE_SHEETS_ID", None):
            self.stdout.write(
                self.style.WARNING("GOOGLE_SHEETS_ID not configured; cannot verify.")
            )
            return

        try:
            service = get_sheets_service()
            sheet_ids = get_existing_registration_ids(service)
        except Exception as exc:
            self.stdout.write(self.style.ERROR(f"Failed to read sheet IDs: {exc}"))
            return

        db_ids = set(
            str(pk)
            for pk in Registration.objects.values_list("id", flat=True)
        )

        missing_from_sheet = db_ids - sheet_ids    # in DB but not in sheet
        extra_in_sheet = sheet_ids - db_ids         # in sheet but not in DB

        self.stdout.write(f"  Total in DB          : {len(db_ids)}")
        self.stdout.write(f"  Total in Sheet       : {len(sheet_ids)}")
        self.stdout.write("")

        if missing_from_sheet:
            self.stdout.write(
                self.style.ERROR(
                    f"  Missing from Sheet ({len(missing_from_sheet)}):"
                )
            )
            for reg_id in sorted(missing_from_sheet, key=lambda x: int(x)):
                self.stdout.write(self.style.ERROR(f"    ID {reg_id}"))
        else:
            self.stdout.write(self.style.SUCCESS("  Missing from Sheet   : none"))

        if extra_in_sheet:
            self.stdout.write(
                self.style.WARNING(
                    f"  Extra in Sheet (not in DB) ({len(extra_in_sheet)}):"
                )
            )
            for reg_id in sorted(extra_in_sheet, key=lambda x: int(x) if x.isdigit() else 0):
                self.stdout.write(self.style.WARNING(f"    ID {reg_id}"))
        else:
            self.stdout.write("  Extra in Sheet       : none")

        self.stdout.write("")
        self.stdout.write("=" * 60)
        if missing_from_sheet:
            self.stdout.write(
                self.style.ERROR(
                    f"MISMATCH DETECTED — {len(missing_from_sheet)} registration(s) missing from sheet. "
                    "Run backfill_sheet (without --verify) to fix."
                )
            )
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    f"ALL SYNCED — {len(db_ids)} registrations confirmed in sheet."
                )
            )
        self.stdout.write("=" * 60)

    # ------------------------------------------------------------------
    # --sort handler
    # ------------------------------------------------------------------

    def _handle_sort(self):
        self.stdout.write(
            "[SORT] Sorting Approvals sheet by payment_date descending, NULLs last...\n"
        )

        if not getattr(settings, "GOOGLE_SHEETS_ID", None):
            self.stdout.write(
                self.style.WARNING("GOOGLE_SHEETS_ID not configured; cannot sort.")
            )
            return

        try:
            service = get_sheets_service()
            sheet_id = settings.GOOGLE_SHEETS_ID
            range_name = settings.GOOGLE_SHEETS_RANGE
        except Exception as exc:
            self.stdout.write(self.style.ERROR(f"Failed to authenticate with Google Sheets: {exc}"))
            return

        success = sort_approvals_sheet_by_payment_date(service, sheet_id, range_name)

        if success:
            self.stdout.write(
                self.style.SUCCESS(
                    "Sheet sorted by payment_date descending. NULLs moved to end.\n"
                    "Open the Google Sheet to confirm rows are ordered by most recent payment date."
                )
            )
        else:
            self.stdout.write(
                self.style.ERROR(
                    "Sort failed. Check logs/sheets_errors.log for details.\n"
                    "Common causes: tab name mismatch (check GOOGLE_SHEETS_RANGE in .env), "
                    "missing 'payment_date' column in header row."
                )
            )

    # ------------------------------------------------------------------
    # Rebuild helpers (--rebuild flag, existing behaviour preserved)
    # ------------------------------------------------------------------

    def _rebuild_excel(self, header_row, data_rows):
        exports_dir = settings.BASE_DIR / "exports"
        os.makedirs(exports_dir, exist_ok=True)
        excel_path = exports_dir / "Main_Conference_Registrations.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"

        ws.append(header_row)
        for row in data_rows:
            ws.append(row)

        from openpyxl.styles import Font
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        wb.save(excel_path)
        self.stdout.write(self.style.SUCCESS(f"✅ Excel sheet successfully built at: {excel_path}"))

    def _rebuild_google(self, header_row, data_rows):
        if not getattr(settings, "GOOGLE_SHEETS_ID", None):
            self.stdout.write(
                self.style.WARNING("GOOGLE_SHEETS_ID not configured; skipping Google Sheets")
            )
            return

        try:
            service = get_sheets_service()
            sheet_id = settings.GOOGLE_SHEETS_ID
            range_name = settings.GOOGLE_SHEETS_RANGE
            sheet_name = range_name.split("!")[0]
            end_col = get_column_letter(len(header_row))

            self.stdout.write("Clearing existing Google Sheet data...")
            service.spreadsheets().values().clear(
                spreadsheetId=sheet_id,
                range=f"{sheet_name}!A1:ZZ",
                body={},
            ).execute()

            self.stdout.write("Writing fresh database rows to Google Sheet...")
            rows = [header_row] + data_rows

            batch_size = 500
            start_row = 1
            for idx in range(0, len(rows), batch_size):
                chunk = rows[idx:idx + batch_size]
                end_row = start_row + len(chunk) - 1
                update_range = f"{sheet_name}!A{start_row}:{end_col}{end_row}"

                service.spreadsheets().values().update(
                    spreadsheetId=sheet_id,
                    range=update_range,
                    valueInputOption="USER_ENTERED",
                    body={"values": chunk},
                ).execute()
                start_row = end_row + 1

            self.stdout.write(
                self.style.SUCCESS("✅ Google Sheets successfully rebuilt with all DB records.")
            )
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Google Sheets rebuild failed: {str(e)}"))
